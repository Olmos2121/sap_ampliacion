"""
preparar_materiales.py
======================
Transforma el Excel de altas de materiales en un Excel listo para
pegar en la app de ampliación SAP.

Uso:
    python preparar_materiales.py <archivo_entrada.xlsx> [<salida.xlsx>]

Si no se especifica salida, genera:
    <nombre_entrada>_SAP_<fecha>.xlsx
"""

import sys
import re
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# Tabla de conversión de IVA: valor normalizado → TAXIM SAP
# ─────────────────────────────────────────────────────────────────────────────
IVA_A_TAXIM = {
    "0.21": "1",
    "21":   "1",
    "0.27": "2",
    "27":   "2",
    "0.105": "3",
    "10.5":  "3",
    "0.025": "4",
    "2.5":   "4",
    "0.05":  "5",
    "5":     "5",
    "0":     "6",
    "exento": "6",
    "no alcanzado": "7",
}

# Grupo de imputación por tipo de material
# Si tiene un solo valor → se completa automáticamente
# Si tiene múltiples → se deja para que el usuario elija
KTGRM_POR_ZMAT = {
    "ZMED":  ["01", "02"],
    "ZSER":  ["04", "05"],
    "ZNOM":  ["03"],
    "ZINS":  ["03"],
    "ZMAT":  ["03"],
    "ZPMA":  ["03"],
    "ZCOM":  ["03"],
    "ZNOA":  ["03"],
    "ZNOV":  ["03"],
}

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def normalizar_iva(valor: str) -> str:
    if not valor or pd.isna(valor):
        return ""
    v = str(valor).strip().lower().replace("%", "").replace(",", ".")
    return IVA_A_TAXIM.get(v, "")


def limpiar_texto(valor, default: str = "") -> str:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return default
    s = str(valor).strip()
    # Colapsar múltiples espacios y saltos de línea
    s = re.sub(r"[\r\n]+", " ", s)
    s = re.sub(r" {2,}", " ", s)
    return s.strip()


def leer_tabla_conversion(ruta: str) -> pd.DataFrame:
    """
    Lee la hoja TABLAS CONV del Excel de conversión y devuelve un
    DataFrame con las columnas:
        ID_CATEGORIA, ZMAT, SPART, XCHPF, CENTRO_SUM, EKGRP
    """
    df = pd.read_excel(ruta, header=1, dtype=str)
    conv = df[["Category2", "Tipo de material SAP", "Sector",
               "Gestion de lote", "Centro sum", "Grupo de compra"]].copy()
    conv.columns = ["ID_CATEGORIA", "ZMAT", "SPART", "XCHPF", "CENTRO_SUM", "EKGRP"]
    conv = conv[conv["ID_CATEGORIA"].str.match(r"^J\d{4}$", na=False)]
    conv = conv.drop_duplicates(subset="ID_CATEGORIA")
    return conv.set_index("ID_CATEGORIA")


def leer_materiales(ruta: str) -> pd.DataFrame:
    """
    Lee la hoja ABM del Excel de altas, descarta filas de TICKET y
    filas sin ID_CATEGORIA válido.
    """
    df = pd.read_excel(ruta, sheet_name="ABM", header=2, dtype=str)
    df = df[
        df["ID_CATEGORIA"].notna() &
        df["ID_CATEGORIA"].str.match(r"^J\d{4}$", na=False)
    ].copy()
    df = df.reset_index(drop=True)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Transformación principal
# ─────────────────────────────────────────────────────────────────────────────
def transformar(df_mats: pd.DataFrame, conv: pd.DataFrame) -> pd.DataFrame:
    """
    Cruza el DataFrame de materiales con la tabla de conversión y
    devuelve el DataFrame de salida listo para la app.
    """
    filas = []

    for _, row in df_mats.iterrows():
        id_cat = str(row.get("ID_CATEGORIA", "")).strip()
        regla  = conv.loc[id_cat] if id_cat in conv.index else None

        # ── Campos que vienen directo del Excel de entrada ──────────────
        nombre_sap  = limpiar_texto(row.get("NOMBRE MATERIAL SAP", ""))
        volumen     = limpiar_texto(row.get("Volumen (CM3)", ""))
        iva_raw     = limpiar_texto(row.get("IVA", ""))
        ecommerce = limpiar_texto(row.get("E-Commerce", ""), default="No especificado").upper()
        if ecommerce == "":
            ecommerce = "No especificado"
        prdha = limpiar_texto(row.get("Unnamed: 4", ""))   # Category3 del árbol
        matkl = id_cat  # Category2 del árbol = ID_CATEGORIA

        # ── Campos derivados de la tabla de conversión ──────────────────
        zmat      = regla["ZMAT"]     if regla is not None else ""
        spart     = regla["SPART"]    if regla is not None else ""
        centro    = regla["CENTRO_SUM"] if regla is not None else ""
        ekgrp     = regla["EKGRP"]    if regla is not None else ""

        # ── TAXIM desde IVA ─────────────────────────────────────────────
        taxim = normalizar_iva(iva_raw)

        # ── KTGRM desde tipo de material ─────────────────────────────────
        opciones_ktgrm = KTGRM_POR_ZMAT.get(zmat, [])
        if len(opciones_ktgrm) == 1:
            ktgrm = opciones_ktgrm[0]
            ktgrm_pendiente = False
        elif len(opciones_ktgrm) > 1:
            ktgrm = " / ".join(opciones_ktgrm)
            ktgrm_pendiente = True
        else:
            ktgrm = ""
            ktgrm_pendiente = False

        # ── Centros a ampliar ────────────────────────────────────────────
        # El centro base viene de CENTRO_SUM.
        # Si E-Commerce = SI → también va a A130.
        centros = [centro] if centro else []
        if ecommerce == "SI" and "A130" not in centros:
            centros.append("A130")
        centros_str = " + ".join(sorted(set(centros)))

        filas.append({
            "MATNR":           "",           # Se completa después (viene de Plex)
            "Tipo material":   zmat,
            "MAKTX":           nombre_sap,
            "TEXTO_LARGO":     nombre_sap,
            "MATKL":           matkl,
            "PRDHA":           prdha,  # Se completa con MATNR después de Plex
            "VOLUM":           volumen,
            "SPART":           spart,
            "EKGRP":           ekgrp,
            "TAXIM":           taxim,
            "E-Commerce":      ecommerce,
            "Centros":         centros_str,
            # Campos de referencia (no van a la app)
            "_IVA_orig":       iva_raw,
            "_ID_CATEGORIA":   id_cat,
            "KTGRM":           ktgrm,
            "_ktgrm_pendiente": ktgrm_pendiente,
        })

    return pd.DataFrame(filas)


# ─────────────────────────────────────────────────────────────────────────────
# Escritura del Excel de salida
# ─────────────────────────────────────────────────────────────────────────────
def escribir_salida(df: pd.DataFrame, ruta_salida: str):
    """
    Genera el Excel de salida con dos hojas:
      - PARA_APP: los campos que se pegan en la app (sin columnas _xxx)
      - REFERENCIA: todas las columnas incluyendo los campos originales
    """
    cols_app = ["MATNR", "Tipo material", "MAKTX", "TEXTO_LARGO",
                "MATKL", "PRDHA", "VOLUM", "SPART", "EKGRP", "TAXIM",
                "E-Commerce", "Centros", "KTGRM"]

    df_app = df[cols_app].copy()
    df_ref = df.copy()
    df_ref.columns = [c.lstrip("_") for c in df_ref.columns]

    wb = openpyxl.Workbook()

    # ── Hoja PARA_APP ──────────────────────────────────────────────────
    ws_app = wb.active
    ws_app.title = "PARA_APP"

    HEADER_FILL  = PatternFill("solid", fgColor="0A6ED1")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    PENDING_FILL = PatternFill("solid", fgColor="FFF2CC")  # amarillo = pendiente
    THIN         = Side(style="thin")
    BORDER       = Border(bottom=THIN)

    # Encabezados
    for col_idx, col_name in enumerate(df_app.columns, 1):
        cell = ws_app.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    ws_app.row_dimensions[1].height = 22

    # Datos
    CHOICE_FILL = PatternFill("solid", fgColor="FFE0B2")

    for row_idx, (_, row_full) in enumerate(df.iterrows(), 2):
        ktgrm_pend = row_full.get("_ktgrm_pendiente", False)
        for col_idx, col_name in enumerate(df_app.columns, 1):
            value = row_full.get(col_name, "")
            cols_no_especificado = {"E-Commerce", "TAXIM", "EKGRP", "SPART"}
            valor_celda = value
            if value in ("", None) or (isinstance(value, float) and pd.isna(value)):
                valor_celda = "No especificado" if col_name in cols_no_especificado else None
            cell = ws_app.cell(row=row_idx, column=col_idx, value=valor_celda)
            if col_name == "MATNR" and not value:
                cell.fill = PENDING_FILL
            elif col_name == "KTGRM" and ktgrm_pend:
                cell.fill = CHOICE_FILL

    # Anchos de columna
    anchos = {
        "MATNR": 15, "Tipo material": 14, "MAKTX": 40, "TEXTO_LARGO": 40,
        "MATKL": 10, "PRDHA": 14, "VOLUM": 10, "SPART": 8,
        "EKGRP": 8,  "TAXIM": 8, "E-Commerce": 17, "Centros": 16,
        "KTGRM": 8,
    }
    for col_idx, col_name in enumerate(df_app.columns, 1):
        ws_app.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = anchos.get(col_name, 14)

    # Freeze first row
    ws_app.freeze_panes = "A2"

    # ── Hoja REFERENCIA ────────────────────────────────────────────────
    ws_ref = wb.create_sheet("REFERENCIA")
    REF_FILL = PatternFill("solid", fgColor="3D5A80")
    REF_FONT = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, col_name in enumerate(df_ref.columns, 1):
        cell = ws_ref.cell(row=1, column=col_idx, value=col_name)
        cell.fill = REF_FILL
        cell.font = REF_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df_ref.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws_ref.cell(row=row_idx, column=col_idx,
                        value=value if value != "" else None)

    for col_idx in range(1, len(df_ref.columns) + 1):
        ws_ref.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = 18

    ws_ref.freeze_panes = "A2"

    wb.save(ruta_salida)
    print(f"✅ Guardado: {ruta_salida}")
    print(f"   {len(df_app)} materiales procesados")

    # Resumen
    sin_matnr = df_app["MATNR"].isna().sum() + (df_app["MATNR"] == "").sum()
    sin_taxim = (df_app["TAXIM"] == "").sum()
    sin_zmat  = (df_app["Tipo material"] == "").sum()

    if sin_matnr:
        print(f"   ⚠️  {sin_matnr} materiales sin MATNR (completar después de Plex)")
    if sin_taxim:
        print(f"   ⚠️  {sin_taxim} materiales sin TAXIM (IVA no reconocido)")
    if sin_zmat:
        print(f"   ⚠️  {sin_zmat} materiales sin tipo de material (ID_CATEGORIA no encontrado en tabla)")


# ─────────────────────────────────────────────────────────────────────────────
# Tabla de conversión embebida (ruta relativa al script)
# ─────────────────────────────────────────────────────────────────────────────
CONV_PATH = Path(__file__).parent / "Conversion_Materiales_SAP.xlsx"


def main():
    if len(sys.argv) < 2:
        print("Uso: python preparar_materiales.py <entrada.xlsx> [salida.xlsx]")
        print()
        print("  entrada.xlsx  Excel de altas de materiales (hoja ABM)")
        print("  salida.xlsx   Nombre del archivo de salida (opcional)")
        sys.exit(1)

    ruta_entrada = sys.argv[1]
    if len(sys.argv) >= 3:
        ruta_salida = sys.argv[2]
    else:
        fecha = datetime.now().strftime("%Y%m%d_%H%M")
        stem  = Path(ruta_entrada).stem
        ruta_salida = str(Path(ruta_entrada).parent / f"{stem}_SAP_{fecha}.xlsx")

    if not Path(ruta_entrada).exists():
        print(f"❌ No se encontró el archivo: {ruta_entrada}")
        sys.exit(1)

    if not CONV_PATH.exists():
        print(f"❌ No se encontró la tabla de conversión: {CONV_PATH}")
        print("   Asegurate de que 'Conversion_Materiales_SAP.xlsx' esté en la misma carpeta que el script.")
        sys.exit(1)

    print(f"📂 Leyendo: {ruta_entrada}")
    conv     = leer_tabla_conversion(str(CONV_PATH))
    df_mats  = leer_materiales(ruta_entrada)
    print(f"   {len(df_mats)} materiales válidos encontrados")

    df_out = transformar(df_mats, conv)
    escribir_salida(df_out, ruta_salida)


if __name__ == "__main__":
    main()