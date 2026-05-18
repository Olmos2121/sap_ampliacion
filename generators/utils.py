"""
utils.py — Helpers compartidos para construir hojas SAP y escribir el Excel.
"""
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Colores de encabezado para cada vista (coinciden con la leyenda del Excel original)
COLORES_VISTA = {
    "Datos básicos":             "FFC000",
    "Datos de centro":           "92D050",
    "Cadenas de distribución":   "00B0F0",
    "Clasificación fiscal":      "7030A0",
    "Datos de previsión":        "FF0000",
    "Lugares de almacenamiento": "00B050",
    "Área planific.nec.":        "ED7D31",
    "Datos de valoración":       "4472C4",
}


def _estilo_encabezado(ws, nombre_hoja: str):
    """Aplica color de fondo, fuente blanca y borde a la fila de encabezado."""
    color = COLORES_VISTA.get(nombre_hoja, "D9D9D9")
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF" if color != "D9D9D9" else "000000")
    thin = Side(style="thin")
    border = Border(bottom=Side(style="medium"))
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _fila_tecnica(ws, campos_tecnicos: list):
    """Agrega la fila 2 con los nombres técnicos SAP (campo SAP)."""
    for col_idx, campo in enumerate(campos_tecnicos, 1):
        cell = ws.cell(row=2, column=col_idx, value=campo)
        cell.font = Font(italic=True, color="595959")
        cell.alignment = Alignment(horizontal="center")


def escribir_excel(hojas: dict) -> bytes:
    """
    Recibe un dict {nombre_hoja: DataFrame} y devuelve bytes de un .xlsx
    con formato similar al original.
    hojas es un OrderedDict para mantener el orden.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja vacía default

    for nombre, df in hojas.items():
        if df is None or df.empty:
            continue
        ws = wb.create_sheet(title=nombre[:31])  # Excel limita a 31 chars

        # Encabezado (fila 1): nombres humanos
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Fila 2: campos técnicos SAP (guardados en df.attrs si existen)
        campos_tecnicos = df.attrs.get("campos_tecnicos", [])
        if campos_tecnicos:
            for col_idx, campo in enumerate(campos_tecnicos, 1):
                cell = ws.cell(row=2, column=col_idx, value=campo)
                cell.font = Font(italic=True, color="595959")
                cell.alignment = Alignment(horizontal="center")
            fila_inicio_datos = 3
        else:
            fila_inicio_datos = 2

        # Datos
        for row_idx, row in enumerate(df.itertuples(index=False), fila_inicio_datos):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value if value != "" else None)

        # Estilos
        _estilo_encabezado(ws, nombre)

        # Ancho de columnas automático
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def construir_datos_basicos(df_input: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    """
    Construye el DataFrame de la hoja 'Datos básicos'.
    df_input debe tener: MATNR, MAKTX, MATKL, PRDHA, VOLUM (y otros variables).
    cfg es el dict del tipo de material desde TIPOS_MATERIAL.
    """
    from config import VISTA_COLUMNAS
    fijos = cfg["valores_fijos_basicos"]
    vistas = cfg["vistas_basicos"]

    rows = []
    for _, mat in df_input.iterrows():
        row = {"Número de producto": mat["MATNR"]}

        # Columnas de vistas (X o vacío)
        for codigo, col_nombre in VISTA_COLUMNAS.items():
            if codigo == "MRP_AREA":
                row[col_nombre] = "X" if "MRP" in vistas else ""
            else:
                row[col_nombre] = "X" if codigo in vistas else ""

        row["Clase de producto"] = fijos.get("MTART", "")
        row["Grupo de productos"] = mat.get("MATKL", "")
        row["Descripción"] = mat.get("MAKTX", "")
        row["Código de idioma"] = fijos.get("SPRAS", "ES")
        row["Partición"] = fijos.get("SPART", "")
        row["Jerarquía de productos"] = mat.get("PRDHA", "")
        row["Indicador Gestión de lotes"] = fijos.get("XCHPF", "")
        row["Grupo de clases de artículo general"] = fijos.get("MTPOS", "")
        row["Volumen"] = mat.get("VOLUM", "")
        row["Unidad de volumen"] = fijos.get("VOLEH", "")
        row["Clave de valor de compras"] = fijos.get("EKWSL", "")
        row["Grupo de transporte"] = fijos.get("TRAGR", "")
        row["Indicador de periodo para caducidad"] = fijos.get("IPRKZ", "")
        row["Caducidad residual mínima"] = fijos.get("MHDRZ", "")
        row["Perfil Serie"] = fijos.get("SERIAL", "")
        row["Texto largo"] = mat.get("TEXTO_LARGO", mat.get("MAKTX", ""))
        row["Bloqueo de material"] = fijos.get("MSTAE", "/")
        rows.append(row)

    df = pd.DataFrame(rows).fillna("")
    df.attrs["campos_tecnicos"] = [
        "MATNR", "", "", "", "", "", "", "", "", "", "MTART", "MATKL",
        "MAKTX", "SPRAS", "SPART", "PRDHA", "XCHPF", "MTPOS", "VOLUM",
        "VOLEH", "EKWSL", "TRAGR", "IPRKZ", "MHDRZ", "SERIAL", "", "MSTAE"
    ]
    return df


def construir_clasificacion_fiscal(df_input: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    """Hoja Clasificación fiscal (una fila por material)."""
    fiscal = cfg.get("fiscal")
    if not fiscal:
        return pd.DataFrame()

    rows = []
    for _, mat in df_input.iterrows():
        rows.append({
            "Número de producto": mat["MATNR"],
            "País/Región": "AR",
            "Clase de impuesto 1": "J1AU",
            "Clasificación fiscal 1": fiscal.get("TAXM1", ""),
            "Clase de impuesto 2": "J901",
            "Clasificación fiscal 2": fiscal.get("TAXM2", "0"),
            "Clase de impuesto 3": "", "Clasificación fiscal 3": "",
            "Clase de impuesto 4": "", "Clasificación fiscal 4": "",
            "Clase de impuesto 5": "", "Clasificación fiscal 5": "",
        })
    df = pd.DataFrame(rows).fillna("")
    df.attrs["campos_tecnicos"] = [
        "MATNR", "ALAND", "TATYP1", "TAXM1", "TATYP2", "TAXM2",
        "TATYP3", "TAXM3", "TATYP4", "TAXM4", "TATYP5", "TAXM5"
    ]
    return df


def construir_datos_prevision(df_input: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    """Hoja Datos de previsión (solo ZMED y ZNOM, centro A130)."""
    if not cfg.get("prevision"):
        return pd.DataFrame()
    rows = []
    for _, mat in df_input.iterrows():
        rows.append({
            "Número de producto": mat["MATNR"],
            "Centro": "A130",
            "Modelo de pronóstico": "J",
        })
    df = pd.DataFrame(rows)
    df.attrs["campos_tecnicos"] = ["MATNR", "WERKS", "PRMOD"]
    return df


def construir_area_planificacion(df_input: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    """Hoja Área planif.nec. (solo ZMED y ZNOM)."""
    if not cfg.get("area_planificacion"):
        return pd.DataFrame()
    rows = []
    for _, mat in df_input.iterrows():
        rows.append({
            "Número de producto": mat["MATNR"],
            "Centro": "A130",
            "Área planif.MRP": "A130_0040",
            "Tipo de MRP": "V1",
            "Planificador de necesidades": "Z01",
            "Grupo MRP": "0000",
            "Punto de reorganización": "10",
            "Cálculo de tamaño de lote": "EX",
            "Tamaño de lote mínimo": "1",
            "Tamaño de lote máximo": "",
            "Tamaño de lote fijo": "",
            "Hora de entrega planificada": "4",
            "Procedimiento de selección de modelo": "2",
            "Indicador de inicialización": "X",
            "Modelo de pronóstico": "J",
            "Límite de alarma": "4",
            "Cantidad de períodos del pasado": "60",
            "Cantidad de períodos de pronóstico": "12",
            "Anular automáticamente modelo pronóstico": "X",
        })
    df = pd.DataFrame(rows).fillna("")
    df.attrs["campos_tecnicos"] = [
        "MATNR", "WERKS", "BERID", "DISMM", "DISPO", "DISGR", "MINBE",
        "DISLS", "BSTMI", "BSTMA", "BSTFE", "PLIFZ", "MODAV", "KZINI",
        "PRMOD", "SIGGR", "PERAN", "ANZPR", "AUTRU"
    ]
    return df
