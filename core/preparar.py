"""
core/preparar.py
Lógica de transformación del Excel de altas al formato de entrada de la app.
"""
import re
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# Tablas de conversión
# ─────────────────────────────────────────────────────────────────────────────
IVA_A_TAXIM = {
    "0.21": "1", "21": "1",
    "0.27": "2", "27": "2",
    "0.105": "3", "10.5": "3",
    "0.025": "4", "2.5": "4",
    "0.05": "5", "5": "5",
    "0": "6", "exento": "6",
    "no alcanzado": "7",
}

KTGRM_POR_ZMAT = {
    "ZMED":  ["01", "02"],
    "ZSER":  ["04", "05"],
    "ZNOM":  ["03"],
    "ZINS":  ["03"],
    "ZMAT":  ["03"],
    "ZPMA":  ["03"],
    "ZCOM":  ["03"],
    "ZNOA":  ["03"],
    "ZNOV":  ["03"],
}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def limpiar_texto(valor, default: str = "") -> str:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return default
    s = str(valor).strip()
    s = re.sub(r"[\r\n]+", " ", s)
    s = re.sub(r" {2,}", " ", s)
    return s.strip() or default


def normalizar_iva(valor: str) -> str:
    if not valor or pd.isna(valor):
        return ""
    v = str(valor).strip().lower().replace("%", "").replace(",", ".")
    return IVA_A_TAXIM.get(v, "")


# ─────────────────────────────────────────────────────────────────────────────
# Lectura
# ─────────────────────────────────────────────────────────────────────────────
def leer_tabla_conversion(archivo_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(archivo_bytes), header=1, dtype=str)
    conv = df[["Category2", "Tipo de material SAP", "Sector",
               "Gestion de lote", "Centro sum", "Grupo de compra"]].copy()
    conv.columns = ["ID_CATEGORIA", "ZMAT", "SPART", "XCHPF", "CENTRO_SUM", "EKGRP"]
    conv = conv[conv["ID_CATEGORIA"].str.match(r"^J\d{4}$", na=False)]
    conv = conv.drop_duplicates(subset="ID_CATEGORIA")
    return conv.set_index("ID_CATEGORIA")


def leer_materiales(archivo_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(archivo_bytes), sheet_name="ABM", header=2, dtype=str)
    df = df[
        df["ID_CATEGORIA"].notna() &
        df["ID_CATEGORIA"].str.match(r"^J\d{4}$", na=False)
    ].copy()
    return df.reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# Transformación
# ─────────────────────────────────────────────────────────────────────────────
def transformar(df_mats: pd.DataFrame, conv: pd.DataFrame) -> pd.DataFrame:
    filas = []

    for _, row in df_mats.iterrows():
        id_cat  = str(row.get("ID_CATEGORIA", "")).strip()
        regla   = conv.loc[id_cat] if id_cat in conv.index else None

        nombre_sap  = limpiar_texto(row.get("NOMBRE MATERIAL SAP", ""))
        volumen     = limpiar_texto(row.get("Volumen (CM3)", ""))
        iva_raw     = limpiar_texto(row.get("IVA", ""))
        ecommerce   = limpiar_texto(row.get("E-Commerce", ""), default="No especificado").upper()
        if ecommerce == "":
            ecommerce = "No especificado"
        trazable_raw = limpiar_texto(row.get("Trazable", ""), default="").upper()
        trazable    = "SI" if trazable_raw == "SI" else "NO"
        prdha       = limpiar_texto(row.get("Unnamed: 4", ""))
        matkl       = id_cat

        zmat    = regla["ZMAT"]      if regla is not None else ""
        spart   = regla["SPART"]     if regla is not None else ""
        centro  = regla["CENTRO_SUM"] if regla is not None else ""
        ekgrp   = regla["EKGRP"]     if regla is not None else ""
        taxim   = normalizar_iva(iva_raw)

        # Centros
        centros = [centro] if centro else []
        if ecommerce == "SI" and "A130" not in centros:
            centros.append("A130")
        centros_str = " + ".join(sorted(set(centros)))

        # KTGRM
        opciones = KTGRM_POR_ZMAT.get(zmat, [])
        if len(opciones) == 1:
            ktgrm, ktgrm_pendiente = opciones[0], False
        elif len(opciones) > 1:
            ktgrm, ktgrm_pendiente = " / ".join(opciones), True
        else:
            ktgrm, ktgrm_pendiente = "", False

        filas.append({
            "MATNR":              "",
            "Tipo material":      zmat,
            "MAKTX":              nombre_sap,
            "TEXTO_LARGO":        nombre_sap,
            "MATKL":              matkl,
            "PRDHA":              prdha,
            "VOLUM":              volumen,
            "SPART":              spart,
            "EKGRP":              ekgrp,
            "TAXIM":              taxim or "No especificado",
            "E-Commerce":         ecommerce,
            "Trazable":           trazable,
            "Centros":            centros_str,
            "KTGRM":              ktgrm,
            "_ktgrm_pendiente":   ktgrm_pendiente,
            "_ID_CATEGORIA":      id_cat,
            "_IVA_orig":          iva_raw,
        })

    return pd.DataFrame(filas)


# ─────────────────────────────────────────────────────────────────────────────
# Escritura
# ─────────────────────────────────────────────────────────────────────────────
def generar_excel(df: pd.DataFrame) -> bytes:
    cols_app = ["MATNR", "Tipo material", "MAKTX", "TEXTO_LARGO",
                "MATKL", "PRDHA", "VOLUM", "SPART", "EKGRP",
                "TAXIM", "E-Commerce", "Trazable", "Centros", "KTGRM"]
    df_app = df[cols_app].copy()

    HEADER_FILL  = PatternFill("solid", fgColor="0A6ED1")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    PENDING_FILL = PatternFill("solid", fgColor="FFF2CC")
    CHOICE_FILL  = PatternFill("solid", fgColor="FFE0B2")
    THIN         = Side(style="thin")
    BORDER       = Border(bottom=THIN)

    anchos = {
        "MATNR": 15, "Tipo material": 14, "MAKTX": 40, "TEXTO_LARGO": 40,
        "MATKL": 10, "PRDHA": 12, "VOLUM": 10, "SPART": 8,
        "EKGRP": 8,  "TAXIM": 8,  "E-Commerce": 14, "Trazable": 10, "Centros": 18, "KTGRM": 10,
    }

    wb = openpyxl.Workbook()

    # ── Hoja PARA_APP ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "PARA_APP"

    for col_idx, col_name in enumerate(df_app.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    cols_no_esp = {"E-Commerce", "TAXIM", "EKGRP", "SPART"}

    for row_idx, (_, row_full) in enumerate(df.iterrows(), 2):
        ktgrm_pend = row_full.get("_ktgrm_pendiente", False)
        for col_idx, col_name in enumerate(df_app.columns, 1):
            value = row_full.get(col_name, "")
            if value is None or (isinstance(value, float) and pd.isna(value)) or value == "":
                valor_celda = "No especificado" if col_name in cols_no_esp else None
            else:
                valor_celda = value
            cell = ws.cell(row=row_idx, column=col_idx, value=valor_celda)
            if col_name == "MATNR" and not value:
                cell.fill = PENDING_FILL
            elif col_name == "KTGRM" and ktgrm_pend:
                cell.fill = CHOICE_FILL

    for col_idx, col_name in enumerate(df_app.columns, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = anchos.get(col_name, 14)

    # ── Hoja REFERENCIA ────────────────────────────────────────────────
    ws_ref = wb.create_sheet("REFERENCIA")
    REF_FILL = PatternFill("solid", fgColor="3D5A80")
    REF_FONT = Font(bold=True, color="FFFFFF", size=10)
    df_ref   = df.copy()
    df_ref.columns = [c.lstrip("_") for c in df_ref.columns]

    for col_idx, col_name in enumerate(df_ref.columns, 1):
        cell = ws_ref.cell(row=1, column=col_idx, value=col_name)
        cell.fill = REF_FILL
        cell.font = REF_FONT
        cell.alignment = Alignment(horizontal="center")
    ws_ref.freeze_panes = "A2"

    for row_idx, row in enumerate(df_ref.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws_ref.cell(row=row_idx, column=col_idx,
                        value=value if value not in ("", None) else None)

    for col_idx in range(1, len(df_ref.columns) + 1):
        ws_ref.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Función principal para llamar desde app.py
# ─────────────────────────────────────────────────────────────────────────────
def procesar(bytes_altas: bytes, bytes_conv: bytes) -> tuple[bytes, int, list]:
    """
    Procesa el Excel de altas y devuelve:
      (bytes_excel_salida, n_materiales, advertencias)
    """
    conv    = leer_tabla_conversion(bytes_conv)
    df_mats = leer_materiales(bytes_altas)
    df_out  = transformar(df_mats, conv)
    excel   = generar_excel(df_out)

    advertencias = []
    sin_taxim = (df_out["TAXIM"] == "No especificado").sum()
    sin_zmat  = (df_out["Tipo material"] == "").sum()
    if sin_taxim:
        advertencias.append(f"{sin_taxim} material(es) con IVA no reconocido → TAXIM = 'No especificado'")
    if sin_zmat:
        advertencias.append(f"{sin_zmat} material(es) con ID_CATEGORIA no encontrado en la tabla de conversión")

    return excel, len(df_out), advertencias